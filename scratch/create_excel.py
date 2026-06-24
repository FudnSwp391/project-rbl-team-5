import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "VTFP-01"
    
    # Enable grid lines visibility
    ws.sheet_view.showGridLines = True
    
    # Fonts
    font_family = "Segoe UI"
    font_title = Font(name=font_family, size=11, bold=True)
    font_bold = Font(name=font_family, size=10, bold=True)
    font_regular = Font(name=font_family, size=10)
    font_italic = Font(name=font_family, size=10, italic=True)
    font_header_vertical = Font(name=font_family, size=9, bold=True, color="FFFFFF")
    font_section_vertical = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    
    # Fills
    fill_dark_blue = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    fill_light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_header_gray = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    # Borders
    border_color = "A6A6A6"
    thin = Side(style='thin', color=border_color)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Alignments
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_vertical_header = Alignment(text_rotation=90, horizontal='center', vertical='center')
    align_vertical_section = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    # Helper to style a range of cells (including borders for merged cells)
    def style_range(ws, cell_range, font=None, fill=None, alignment=None, border=None):
        for row in ws[cell_range]:
            for cell in row:
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if alignment:
                    cell.alignment = alignment
                if border:
                    cell.border = border

    # --- 1. Metadata Block (Rows 1 to 4) ---
    ws.merge_cells("B1:D1")
    ws.merge_cells("F1:Q1")
    ws.merge_cells("B2:D2")
    ws.merge_cells("F2:Q2")
    ws.merge_cells("B3:D3")
    ws.merge_cells("F3:Q3")
    ws.merge_cells("B4:Q4")
    
    metadata = [
        ("A1", "Function Code", "VTFP-01", "E1", "Function Name", "VTFP-01_Login via Google"),
        ("A2", "Created By", "MyPTT", "E2", "Executed By", "MyPTT"),
        ("A3", "Lines of code", 30, "E3", "Lack of test cases", -7),
        ("A4", "Test requirement", "<Brief description about requirements which are tested in this function>", "", "", "")
    ]
    
    for row_idx, (c1, l1, v1, c2, l2, v2) in enumerate(metadata, start=1):
        ws[c1] = l1
        ws[f"B{row_idx}"] = v1
        if c2:
            ws[c2] = l2
            ws[f"F{row_idx}"] = v2
            
    # Apply styling to Metadata Block
    for r in range(1, 5):
        ws.row_dimensions[r].height = 20
        ws[f"A{r}"].font = font_bold
        ws[f"A{r}"].fill = fill_light_gray
        ws[f"A{r}"].border = border_all
        ws[f"A{r}"].alignment = align_left
        
        if r < 4:
            ws[f"E{r}"].font = font_bold
            ws[f"E{r}"].fill = fill_light_gray
            ws[f"E{r}"].border = border_all
            ws[f"E{r}"].alignment = align_left
            
        for col in ["B", "C", "D"]:
            cell = ws[f"{col}{r}"]
            cell.border = border_all
            cell.font = font_italic if r == 4 else font_regular
            cell.alignment = align_left
        for col in [get_column_letter(x) for x in range(6, 18)]: # F to Q
            cell = ws[f"{col}{r}"]
            cell.border = border_all
            cell.font = font_italic if r == 4 else font_regular
            cell.alignment = align_left
            
    ws["F3"] = "=G6-22"

    # --- 2. Stats Block (Rows 5 to 6) ---
    ws.merge_cells("D5:F5")
    ws["A5"] = "Passed"
    ws["B5"] = "Failed"
    ws["C5"] = "Untested"
    ws["D5"] = "N/A/B"
    ws["G5"] = "Total Test Cases"
    
    ws["A6"] = '=COUNTIF(C38:Q38, "P")'
    ws["B6"] = '=COUNTIF(C38:Q38, "F")'
    ws["C6"] = '=COUNTIF(C38:Q38, "")'
    ws["D6"] = '=COUNTIF(C37:Q37, "N")'
    ws["E6"] = '=COUNTIF(C37:Q37, "A")'
    ws["F6"] = '=COUNTIF(C37:Q37, "B")'
    ws["G6"] = '=COUNTA(C7:Q7)'
    
    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 20
    
    for col_let in ["A", "B", "C", "D", "E", "F", "G"]:
        cell = ws[f"{col_let}5"]
        cell.font = font_bold
        cell.fill = fill_header_gray
        cell.alignment = align_center
        cell.border = border_all
        
    for col_let in ["A", "B", "C", "D", "E", "F", "G"]:
        cell = ws[f"{col_let}6"]
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border_all

    # --- 3. Matrix Header (Row 7) ---
    ws.row_dimensions[7].height = 70
    ws.merge_cells("A7:B7")
    style_range(ws, "A7:B7", fill=fill_dark_blue, border=border_all)
    
    for i in range(1, 16):
        col_let = get_column_letter(2 + i)
        ws[f"{col_let}7"] = f"UTCID{i:02d}"
        cell = ws[f"{col_let}7"]
        cell.font = font_header_vertical
        cell.fill = fill_dark_blue
        cell.alignment = align_vertical_header
        cell.border = border_all

    # --- 4. Matrix Rows ---
    row_data = [
        # (Row, Section, SubSection/Label, IsSubHeader)
        # Condition Section (8 to 19)
        (8, "Condition", "Precondition", True),
        (9, "Condition", "Can connect with Google's OAuth service", False),
        (10, "Condition", "Cannot connect with Google's OAuth service", False),
        (11, "Condition", "User exists in database", False),
        (12, "Condition", "User does not exist (first-time login)", False),
        (13, "Condition", "User is banned or disabled", False),
        (14, "Condition", "Input (Google IdToken)", True),
        (15, "Condition", "Valid token: eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.ey", False),
        (16, "Condition", "Invalid / expired token: eyFakeExpired123", False),
        (17, "Condition", '"" or null', False),
        (18, "Condition", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.error", False),
        (19, "Condition", "", False), # Blank Row
        
        # Confirm Section (20 to 36)
        (20, "Confirm", "Return", True),
        (21, "Confirm", "200 OK (Login successful)", False),
        (22, "Confirm", "201 Created (New user created)", False),
        (23, "Confirm", "400 Bad Request (Empty or invalid token)", False),
        (24, "Confirm", "401 Unauthorized (Banned or expired token)", False),
        (25, "Confirm", "500 Internal Server Error", False),
        (26, "Confirm", "", False), # Blank Row
        (27, "Confirm", "Exception", True),
        (28, "Confirm", "None", False),
        (29, "Confirm", "AuthException (Invalid credentials)", False),
        (30, "Confirm", "System.Exception (Unhandled server error)", False),
        (31, "Confirm", "", False), # Blank Row
        (32, "Confirm", "Log message", True),
        (33, "Confirm", '"User logged in successfully via Google."', False),
        (34, "Confirm", '"New user created via Google login."', False),
        (35, "Confirm", '"Invalid or expired Google token."', False),
        (36, "Confirm", '"Unexpected error during Google authentication."', False),
        
        # Result Section (37 to 40)
        (37, "Result", "Type(N : Normal, A : Abnormal, B : Boundary)", False),
        (38, "Result", "Passed/Failed", False),
        (39, "Result", "Executed Date", False),
        (40, "Result", "Defect ID", False)
    ]
    
    # Write descriptions and subheaders
    for r_idx, sec, desc, is_sub in row_data:
        ws.row_dimensions[r_idx].height = 20
        ws[f"B{r_idx}"] = desc
        cell = ws[f"B{r_idx}"]
        cell.border = border_all
        
        if is_sub:
            cell.font = font_bold
            cell.fill = fill_light_gray
            cell.alignment = align_left
        else:
            if sec == "Result":
                cell.font = font_bold
                cell.fill = fill_light_gray
                cell.alignment = align_left
            elif desc == "":
                # Blank separator rows inside sections
                cell.value = None
            else:
                cell.font = font_regular
                cell.alignment = align_left
                
    # Merge and style column A sections
    ws.merge_cells("A8:A19")
    ws["A8"] = "C\no\nn\nd\ni\nt\ni\no\nn"
    style_range(ws, "A8:A19", font=font_section_vertical, fill=fill_dark_blue, alignment=align_vertical_section, border=border_all)
    
    ws.merge_cells("A20:A36")
    ws["A20"] = "C\no\nn\nf\ni\nr\nm"
    style_range(ws, "A20:A36", font=font_section_vertical, fill=fill_dark_blue, alignment=align_vertical_section, border=border_all)
    
    ws.merge_cells("A37:A40")
    ws["A37"] = "R\ne\ns\nu\nl\nt"
    style_range(ws, "A37:A40", font=font_section_vertical, fill=fill_dark_blue, alignment=align_vertical_section, border=border_all)
    
    # --- 5. Fill Matrix Intersection Data (C to Q) ---
    test_cases_data = {
        'C': { # UTCID01
            9: 'O', 11: 'O', 15: 'O', 21: 'O', 28: 'O', 33: 'O',
            37: 'N', 38: 'P', 39: 1
        },
        'D': { # UTCID02
            9: 'O', 12: 'O', 15: 'O', 21: 'O', 22: 'O', 28: 'O', 34: 'O',
            37: 'N', 38: 'P', 39: 1
        },
        'E': { # UTCID03
            9: 'O', 13: 'O', 15: 'O', 24: 'O', 29: 'O', 35: 'O',
            37: 'A', 38: 'P', 39: 1
        },
        'F': { # UTCID04
            16: 'O', 25: 'O', 29: 'O', 35: 'O',
            37: 'A', 38: 'P', 39: 1
        },
        'G': { # UTCID05
            10: 'O', 16: 'O', 23: 'O', 29: 'O', 35: 'O',
            37: 'A', 38: 'P', 39: 1
        },
        'H': { # UTCID06
            17: 'O', 23: 'O', 29: 'O', 35: 'O',
            37: 'A', 38: 'P', 39: 1
        },
        'I': { # UTCID07
            18: 'O', 25: 'O', 30: 'O', 36: 'O',
            37: 'A', 38: 'P', 39: 1
        },
        'J': { # UTCID08
            10: 'O', 15: 'O', 25: 'O', 30: 'O', 36: 'O',
            37: 'A', 38: 'P', 39: 1
        },
        'K': { # UTCID09
            10: 'O', 12: 'O', 15: 'O', 25: 'O', 30: 'O', 36: 'O',
            37: 'A', 38: 'P', 39: 1
        },
        'L': { # UTCID10
            10: 'O', 17: 'O', 23: 'O', 29: 'O', 35: 'O',
            37: 'A', 38: 'P', 39: 1
        },
        'M': { # UTCID11
            10: 'O', 18: 'O', 25: 'O', 30: 'O', 36: 'O',
            37: 'A', 38: 'P', 39: 1
        },
        'N': { # UTCID12
            9: 'O', 12: 'O', 16: 'O', 23: 'O', 29: 'O', 35: 'O',
            37: 'A', 38: 'P', 39: 1
        },
        'O': { # UTCID13
            9: 'O', 13: 'O', 16: 'O', 23: 'O', 29: 'O', 35: 'O',
            37: 'A', 38: 'P', 39: 1
        },
        'P': { # UTCID14
            9: 'O', 11: 'O', 17: 'O', 23: 'O', 29: 'O', 35: 'O',
            37: 'A', 38: 'P', 39: 1
        },
        'Q': { # UTCID15
            9: 'O', 12: 'O', 17: 'O', 23: 'O', 29: 'O', 35: 'O',
            37: 'A', 38: 'P', 39: 1
        }
    }
    
    # Prepopulate grid with cells, borders and background colors
    for col_idx in range(3, 18):
        col_let = get_column_letter(col_idx)
        for r_idx in range(8, 41):
            cell = ws[f"{col_let}{r_idx}"]
            cell.border = border_all
            cell.alignment = align_center
            
            # Check if this row is a subheader in column B or a blank separator row
            is_sub_or_blank = False
            for r_num, _, _, is_s in row_data:
                if r_num == r_idx:
                    if is_s or ws[f"B{r_idx}"].value is None:
                        is_sub_or_blank = True
                    break
            
            if is_sub_or_blank:
                cell.fill = fill_light_gray
                continue
                
            val = test_cases_data[col_let].get(r_idx, '')
            if val != '':
                cell.value = val
                if val == 'O':
                    cell.font = font_bold
                elif r_idx in [37, 38]: # Type or Passed/Failed
                    cell.font = font_bold
                else:
                    cell.font = font_regular
            else:
                cell.value = None
                
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    for col_idx in range(3, 18):
        col_let = get_column_letter(col_idx)
        ws.column_dimensions[col_let].width = 9
        
    # Save the file
    output_filename = "VTFP-01_Login_via_Google_TestCases.xlsx"
    workspace_path = r"c:\Users\Huynh Huy\Documents\Techcycle master\project-rbl-team-5"
    full_path = os.path.join(workspace_path, output_filename)
    wb.save(full_path)
    print(f"Excel file updated successfully at: {full_path}")

if __name__ == "__main__":
    create_excel_sheet()
